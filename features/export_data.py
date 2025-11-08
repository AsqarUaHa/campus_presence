# ============================================
# FILE: features/export_data.py
# ============================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import io
import logging

from config import TIMEZONE
from database.db_manager import get_db

logger = logging.getLogger(__name__)


def get_local_time():
    """Получить текущее локальное время"""
    return datetime.now(TIMEZONE)


async def export_presence_data(update, context, period='today', event_id=None):
    """Экспорт данных о присутствии в Excel"""
    query = update.callback_query
    await query.answer("⏳ Генерация файла...")
    
    # Определяем период
    today = get_local_time().date()
    
    if period == 'today':
        start_date = today
        end_date = today
        period_name = "Сегодня"
    elif period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
        period_name = "За неделю"
    elif period == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
        period_name = "За месяц"
    elif period == 'event' and event_id:
        # Экспорт по мероприятию
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT name, start_time, end_time
                FROM events
                WHERE id = %s
            ''', (event_id,))
            event = cursor.fetchone()
            
            if not event:
                await query.message.reply_text("❌ Мероприятие не найдено.")
                return
            
            period_name = event['name']
    
    # Получаем данные
    with get_db() as conn:
        cursor = conn.cursor()
        
        if event_id:
            cursor.execute('''
                SELECT 
                    u.first_name,
                    u.last_name,
                    u.username,
                    u.team_role,
                    u.phone_number,
                    u.birth_date,
                    p.check_in_time,
                    p.check_out_time,
                    p.date
                FROM presence p
                JOIN users u ON p.user_id = u.user_id
                WHERE p.event_id = %s
                ORDER BY p.check_in_time
            ''', (event_id,))
        else:
            cursor.execute('''
                SELECT 
                    u.first_name,
                    u.last_name,
                    u.username,
                    u.team_role,
                    u.phone_number,
                    u.birth_date,
                    p.check_in_time,
                    p.check_out_time,
                    p.date
                FROM presence p
                JOIN users u ON p.user_id = u.user_id
                WHERE p.date BETWEEN %s AND %s
                  AND p.status = 'in_campus'
                ORDER BY p.check_in_time DESC
            ''', (start_date, end_date))
        
        data = cursor.fetchall()
    
    if not data:
        await query.message.reply_text(
            f"📊 Нет данных за период: {period_name}"
        )
        return
    
    # Создаём Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Присутствие"
    
    # Заголовок
    headers = [
        "Имя", "Фамилия", "Username", "Команда/Роль",
        "Телефон", "Дата рождения", "Дата",
        "Время прибытия", "Время ухода", "Длительность (ч)"
    ]
    
    # Стиль заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for row_num, record in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=record['first_name'])
        ws.cell(row=row_num, column=2, value=record['last_name'])
        ws.cell(row=row_num, column=3, value=record['username'] or "")
        ws.cell(row=row_num, column=4, value=record['team_role'] or "")
        ws.cell(row=row_num, column=5, value=record['phone_number'] or "")
        ws.cell(row=row_num, column=6, value=record['birth_date'].strftime('%d.%m.%Y') if record['birth_date'] else "")
        ws.cell(row=row_num, column=7, value=record['date'].strftime('%d.%m.%Y'))
        
        # Время прибытия
        check_in = record['check_in_time']
        if check_in.tzinfo is None:
            check_in = check_in.replace(tzinfo=TIMEZONE)
        local_check_in = check_in.astimezone(TIMEZONE)
        ws.cell(row=row_num, column=8, value=local_check_in.strftime('%H:%M'))
        
        # Время ухода и длительность
        if record['check_out_time']:
            check_out = record['check_out_time']
            if check_out.tzinfo is None:
                check_out = check_out.replace(tzinfo=TIMEZONE)
            local_check_out = check_out.astimezone(TIMEZONE)
            ws.cell(row=row_num, column=9, value=local_check_out.strftime('%H:%M'))
            
            duration = (local_check_out - local_check_in).total_seconds() / 3600
            ws.cell(row=row_num, column=10, value=round(duration, 2))
        else:
            ws.cell(row=row_num, column=9, value="Не ушел")
            ws.cell(row=row_num, column=10, value="")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Отправляем файл
    filename = f"presence_{period}_{get_local_time().strftime('%Y%m%d_%H%M')}.xlsx"
    
    await query.message.reply_document(
        document=excel_file,
        filename=filename,
        caption=f"📊 Экспорт данных: {period_name}\n"
                f"Всего записей: {len(data)}"
    )
